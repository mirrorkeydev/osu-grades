from typing import Dict, List, DefaultDict
from collections import defaultdict
import openpyxl
import csv
import re
import numpy

def translate_sheet_name(name: str) -> str:
  if (match := re.match(r"([A-Z]{1,2})(\d{2})", name)):
    term, year = match.groups()
    if term == "FA" or term == "F":
      return "fall20" + year
    elif term == "WI":
     return "winter20" + year
    elif term == "SP":
     return "spring20" + year
    elif term == "SU":
     return "summer20" + year
  raise Exception(f"unable to translate invalid sheet name {name}")

class Grade:
  def __init__(self, value: str, student_count: int, percent: float) -> None:
    self.value = value
    self.student_count = student_count
    self.percent = percent
  def __str__(self) -> str:
    return f"Grade {self.value}: {self.student_count} students, or {self.percent}%"

class Course:
  def __init__(self, code: str, num: str, term: str, num_students: int, gpa: float) -> None:
    self.code = code # eg. CS
    self.num = num # eg. 160H
    self.term = term
    self.instructor = None # unknown
    self.num_students = num_students
    self.gpa = gpa
    self.grades: List[Grade] = []
  def __str__(self) -> str:
    s = f"{self.code} {self.num} in {self.term} taught by '{self.instructor}', {self.num_students} student(s) and average gpa {self.gpa}\n"
    for grade in self.grades:
      s = s + "\t" + str(grade) + "\n"
    return s

class InstructorCourse:
  def __init__(self, instructor_name: str, course: str) -> None:
    self.instructor_name = instructor_name
    self.course = course
    self.averages = []
    self.terms = []
  def __str__(self) -> str:
    return (f"{self.course} taught by '{self.instructor_name}' awards average GPA {'%.2f'%(numpy.mean(self.averages))} across "
    + f"all terms (n={len(self.terms)})")

courses: List[Course] = []
courseslookup: DefaultDict[str, List[Course]] = defaultdict(list)

wb = openpyxl.load_workbook(filename='data.xlsx', data_only=True)
sheet_names = wb.sheetnames
for sheet_name in sheet_names:
  sheet = wb[sheet_name]

  for row in sheet.iter_rows(values_only=True):
    if row[0] is None:
      continue
    if (match := re.match(r"^Course: (\S+) ([\S]+)  Student Total: (\d+) .+ GPA: ([\d.]+)", row[0])):
      num, code, num_students, avg_gpa = match.groups()
      courses.append(Course(num, code, translate_sheet_name(sheet_name), int(num_students), float(avg_gpa)))
    elif re.match(r"[ \w+]+: (\S+)", row[0]):
      c = courses[-1]
      value = re.match(r"[ \w+]+: (\S+)", row[0]).groups()[0]
      student_count = re.match(r"[ \w+]+: (\S+)", row[2]).groups()[0]
      percent = re.match(r"[ \w+]+: (\S+)", row[3]).groups()[0]
      c.grades.append(Grade(value, student_count, percent))

instructors = {}
with open('all-engineering-instructors-by-term.csv') as f:
  reader = csv.reader(f, delimiter=",", quotechar='"')
  for row in reader:
    instructors[f"{row[0]}{row[1]}{row[2]}"] = row[3]

for course in courses:
  if f"{course.term}{course.code}{course.num}" in instructors:
    course.instructor = instructors[f"{course.term}{course.code}{course.num}"]
  # if course.code == "CS":
  #   print(str(course))
  courseslookup[f"{course.code}{course.num}"].append(course)

while True:
  course_input = input("\nEnter a course (e.g. CS161): ")
  if course_input == "q":
    break
  ics: Dict[str, InstructorCourse] = {}
  for course in courseslookup[course_input]:
    ic = None
    if f"{course.instructor}{course.num}{course.num}" in ics:
      ic = ics[f"{course.instructor}{course.num}{course.num}"]
    else:
      ic = InstructorCourse(course.instructor, f"{course.code}{course.num}")
      ics[f"{course.instructor}{course.num}{course.num}"] = ic
    ic.averages.append(course.gpa)
    ic.terms.append(course.term)

  for _, value in ics.items():
    print(str(value))
    
    


